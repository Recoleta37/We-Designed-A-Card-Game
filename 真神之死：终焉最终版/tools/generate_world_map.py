from pathlib import Path
import math
import random

import openpyxl
from PIL import Image, ImageDraw, ImageFilter, ImageFont


ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "\u6c99\u6f20-\u7cbe\u7075-\u5929\u4f7f-\u9b54\u65cf\u5206\u5e03\u8868.xlsx"
OUT_DIR = ROOT / "WORLD_PVE_\u6781\u9650\u6d4b\u8bd5\u7248" / "assets" / "maps"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT = OUT_DIR / "world_map_from_distribution.png"


def font(size, bold=False):
    candidates = [
        "C:/Windows/Fonts/msyhbd.ttc" if bold else "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/simsun.ttc",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size)
        except OSError:
            continue
    return ImageFont.load_default()


def center_text(draw, text, box, fnt, fill):
    bbox = draw.textbbox((0, 0), text, font=fnt)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    x = box[0] + (box[2] - box[0] - tw) / 2
    y = box[1] + (box[3] - box[1] - th) / 2
    draw.text((x, y), text, font=fnt, fill=fill)


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows, cols = ws.max_row, ws.max_column
    values = [[ws.cell(r, c).value or "" for c in range(1, cols + 1)] for r in range(1, rows + 1)]

    cell = 42
    margin_l, margin_t, margin_r, margin_b = 150, 128, 260, 118
    width = margin_l + cols * cell + margin_r
    height = margin_t + rows * cell + margin_b

    palette = {
        "\u6c99\u6f20": "#d9b56c",
        "\u7cbe\u7075": "#4f8d58",
        "\u4e0d\u6b7b\u5723\u5730": "#70876f",
        "\u5929\u4f7f": "#e9e2c9",
        "\u5723\u6cb3": "#74c8df",
        "\u6492\u51b7": "#efe8d5",
        "\u5438\u8840\u9b3c": "#873348",
        "\u4f0a\u83b2\u5821\u5792": "#542133",
        "\u4eba\u7c7b": "#8bae67",
        "\u738b\u57ce": "#bea565",
        "\u6d77\u6d0b": "#3c7fa6",
        "\u9b54\u65cf": "#52365f",
        "\u9b54\u90fd": "#2b163b",
        "\u5f00\u59cb\u70b9": "#f3d356",
        "\u8fdc\u53e4\u57ce\u5e02": "#b38b58",
        "": "#f1e6d1",
    }

    image = Image.new("RGB", (width, height), "#efe1c2")
    draw = ImageDraw.Draw(image)
    random.seed(9)

    for _ in range(22000):
        x, y = random.randrange(width), random.randrange(height)
        base = random.choice([(245, 229, 193), (230, 207, 162), (255, 241, 211), (214, 189, 139)])
        alpha = random.randint(7, 20)
        old = image.getpixel((x, y))
        image.putpixel((x, y), tuple((old[i] * (255 - alpha) + base[i] * alpha) // 255 for i in range(3)))

    font_title = font(38, True)
    font_sub = font(18)
    font_region = font(29, True)
    font_legend = font(17)
    font_landmark = font(18, True)
    font_tiny = font(12)

    draw.rounded_rectangle([36, 28, width - 36, 92], radius=18, fill="#473624", outline="#b58a49", width=3)
    center_text(draw, "WORLD Alpha \u5730\u7406\u5206\u5e03\u56fe", [36, 25, width - 36, 72], font_title, "#f8df99")
    center_text(
        draw,
        "\u6839\u636e\u300c\u6c99\u6f20-\u7cbe\u7075-\u5929\u4f7f-\u9b54\u65cf\u5206\u5e03\u8868.xlsx\u300d\u751f\u6210\u7684\u4e16\u754c\u5730\u56fe\u7d20\u6750",
        [36, 63, width - 36, 93],
        font_sub,
        "#ead4a3",
    )

    x0, y0 = margin_l, margin_t
    positions = {}
    for r in range(rows):
        for c in range(cols):
            value = str(values[r][c]).strip()
            if value:
                positions.setdefault(value, []).append((x0 + c * cell + cell / 2, y0 + r * cell + cell / 2, r, c))

    draw.rounded_rectangle(
        [x0 - 18, y0 - 18, x0 + cols * cell + 18, y0 + rows * cell + 18],
        radius=22,
        fill="#c6a86a",
        outline="#8a6535",
        width=2,
    )
    draw.rounded_rectangle(
        [x0 - 10, y0 - 10, x0 + cols * cell + 10, y0 + rows * cell + 10],
        radius=18,
        fill="#ecd8ae",
        outline="#644827",
        width=2,
    )

    for r in range(rows):
        for c in range(cols):
            value = str(values[r][c]).strip()
            color = palette.get(value, "#e6d0a0")
            x, y = x0 + c * cell, y0 + r * cell
            draw.rounded_rectangle([x, y, x + cell + 1, y + cell + 1], radius=8, fill=color, outline="#6f512d", width=1)

            if value in ("\u6d77\u6d0b", "\u5723\u6cb3"):
                wave = "#bceefe" if value == "\u5723\u6cb3" else "#9dcbe0"
                for k in range(2):
                    yy = y + 12 + k * 15
                    draw.arc([x + 5, yy - 5, x + 20, yy + 7], 180, 360, fill=wave, width=1)
                    draw.arc([x + 19, yy - 5, x + 34, yy + 7], 180, 360, fill=wave, width=1)
            elif value == "\u6c99\u6f20":
                draw.arc([x + 6, y + 18, x + 30, y + 34], 190, 340, fill="#bc9048", width=1)
                draw.arc([x + 18, y + 10, x + 40, y + 25], 190, 340, fill="#bc9048", width=1)
            elif value == "\u7cbe\u7075":
                draw.ellipse([x + 13, y + 8, x + 29, y + 24], fill="#2f6539")
                draw.rectangle([x + 20, y + 22, x + 23, y + 32], fill="#5b3f24")
            elif value == "\u9b54\u65cf":
                draw.polygon([(x + 5, y + 34), (x + 16, y + 13), (x + 27, y + 34)], fill="#2c1d35")
                draw.polygon([(x + 21, y + 34), (x + 32, y + 12), (x + 42, y + 34)], fill="#392242")
            elif value == "\u5929\u4f7f":
                draw.line([x + 21, y + 9, x + 21, y + 31], fill="#fff8db", width=2)
                draw.line([x + 12, y + 20, x + 30, y + 20], fill="#fff8db", width=2)
                draw.ellipse([x + 15, y + 4, x + 27, y + 16], outline="#fff8db", width=2)
            elif value == "\u4eba\u7c7b":
                draw.rectangle([x + 11, y + 19, x + 31, y + 33], fill="#6d7e52")
                draw.polygon([(x + 9, y + 19), (x + 21, y + 9), (x + 33, y + 19)], fill="#a48548")
            elif value in ("\u5438\u8840\u9b3c", "\u4f0a\u83b2\u5821\u5792"):
                draw.rectangle([x + 10, y + 18, x + 32, y + 34], fill="#351524")
                draw.rectangle([x + 14, y + 10, x + 18, y + 18], fill="#351524")
                draw.rectangle([x + 24, y + 10, x + 28, y + 18], fill="#351524")

    for r in range(rows):
        for c in range(cols):
            value = str(values[r][c]).strip()
            x, y = x0 + c * cell, y0 + r * cell
            if c + 1 < cols and str(values[r][c + 1]).strip() != value:
                draw.line([x + cell, y + 3, x + cell, y + cell - 3], fill="#5f4828", width=2)
            if r + 1 < rows and str(values[r + 1][c]).strip() != value:
                draw.line([x + 3, y + cell, x + cell - 3, y + cell], fill="#5f4828", width=2)

    def centroid(name):
        pts = positions.get(name, [])
        if not pts:
            return None
        return (sum(p[0] for p in pts) / len(pts), sum(p[1] for p in pts) / len(pts))

    route_names = ["\u5f00\u59cb\u70b9", "\u738b\u57ce", "\u4f0a\u83b2\u5821\u5792", "\u4e0d\u6b7b\u5723\u5730", "\u6492\u51b7", "\u9b54\u90fd"]
    route = [centroid(name) for name in route_names if centroid(name)]
    for i in range(len(route) - 1):
        x1, y1 = route[i]
        x2, y2 = route[i + 1]
        steps = int(max(abs(x2 - x1), abs(y2 - y1)) // 14)
        for s in range(steps + 1):
            t = s / max(1, steps)
            x = x1 + (x2 - x1) * t
            y = y1 + (y2 - y1) * t
            draw.ellipse([x - 3, y - 3, x + 3, y + 3], fill="#3d2616")

    landmarks = ["\u5f00\u59cb\u70b9", "\u738b\u57ce", "\u4f0a\u83b2\u5821\u5792", "\u4e0d\u6b7b\u5723\u5730", "\u8fdc\u53e4\u57ce\u5e02", "\u6492\u51b7", "\u9b54\u90fd"]
    for name in landmarks:
        point = centroid(name)
        if not point:
            continue
        x, y = point
        if name == "\u5f00\u59cb\u70b9":
            draw.ellipse([x - 17, y - 17, x + 17, y + 17], fill="#fff1a5", outline="#5c4300", width=3)
            draw.polygon([(x, y - 12), (x + 7, y + 9), (x - 10, y - 2), (x + 10, y - 2), (x - 7, y + 9)], fill="#c4572a")
        elif name in ("\u738b\u57ce", "\u6492\u51b7"):
            draw.rectangle([x - 15, y - 4, x + 15, y + 15], fill="#d5bd73", outline="#3d2c19", width=2)
            draw.rectangle([x - 12, y - 15, x - 5, y - 4], fill="#d5bd73", outline="#3d2c19")
            draw.rectangle([x + 5, y - 15, x + 12, y - 4], fill="#d5bd73", outline="#3d2c19")
        elif name in ("\u4f0a\u83b2\u5821\u5792", "\u9b54\u90fd"):
            draw.rectangle([x - 17, y - 1, x + 17, y + 17], fill="#241327", outline="#f0c48f", width=2)
            draw.rectangle([x - 14, y - 17, x - 7, y - 1], fill="#241327", outline="#f0c48f")
            draw.rectangle([x + 7, y - 17, x + 14, y - 1], fill="#241327", outline="#f0c48f")
        elif name == "\u4e0d\u6b7b\u5723\u5730":
            draw.ellipse([x - 16, y - 16, x + 16, y + 16], fill="#d6decf", outline="#354532", width=2)
            draw.line([x, y - 12, x, y + 13], fill="#354532", width=3)
            draw.line([x - 9, y - 2, x + 9, y - 2], fill="#354532", width=3)
        elif name == "\u8fdc\u53e4\u57ce\u5e02":
            draw.polygon([(x - 18, y + 15), (x, y - 16), (x + 18, y + 15)], fill="#c29b60", outline="#563d22")
            draw.rectangle([x - 6, y + 1, x + 6, y + 15], fill="#886136")

        bbox = draw.textbbox((0, 0), name, font=font_landmark)
        tw = bbox[2] - bbox[0]
        draw.rounded_rectangle([x - tw / 2 - 8, y + 20, x + tw / 2 + 8, y + 44], radius=7, fill="#fff4d6", outline="#6a4d2c")
        draw.text((x - tw / 2, y + 22), name, font=font_landmark, fill="#3b2a17")

    region_labels = [
        ("\u5927\u6c99\u6f20", "\u6c99\u6f20", (-18, -4)),
        ("\u7cbe\u7075\u5723\u5730", "\u7cbe\u7075", (-18, 0)),
        ("\u5929\u4f7f\u56fd\u5ea6", "\u5929\u4f7f", (30, -15)),
        ("\u7329\u7ea2\u5e73\u539f", "\u5438\u8840\u9b3c", (-20, 35)),
        ("\u4eba\u7c7b\u8bf8\u56fd", "\u4eba\u7c7b", (-26, 12)),
        ("\u65e0\u5c3d\u6d77", "\u6d77\u6d0b", (4, 0)),
        ("\u9b54\u5883", "\u9b54\u65cf", (28, 0)),
    ]
    for label, key, offset in region_labels:
        pts = positions.get(key, [])
        if not pts:
            continue
        x = sum(p[0] for p in pts) / len(pts) + offset[0]
        y = sum(p[1] for p in pts) / len(pts) + offset[1]
        bbox = draw.textbbox((0, 0), label, font=font_region)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        bright = key in ("\u9b54\u65cf", "\u5438\u8840\u9b3c", "\u6d77\u6d0b")
        draw.text((x - tw / 2 + 2, y - th / 2 + 2), label, font=font_region, fill="#3c2a16")
        draw.text((x - tw / 2, y - th / 2), label, font=font_region, fill="#fff5d6" if bright else "#3c2b17")

    for c in range(cols):
        if c % 3 == 0:
            draw.text((x0 + c * cell + 14, y0 - 28), str(c + 1), font=font_tiny, fill="#6b512f")
    for r in range(rows):
        if r % 4 == 0:
            draw.text((x0 - 34, y0 + r * cell + 12), str(r + 1), font=font_tiny, fill="#6b512f")

    cx, cy = width - 135, 172
    draw.ellipse([cx - 45, cy - 45, cx + 45, cy + 45], outline="#6c4b28", width=2)
    draw.polygon([(cx, cy - 48), (cx + 10, cy - 5), (cx, cy), (cx - 10, cy - 5)], fill="#493018")
    draw.polygon([(cx, cy + 42), (cx + 8, cy + 6), (cx, cy), (cx - 8, cy + 6)], fill="#b68b4e")
    draw.polygon([(cx - 42, cy), (cx - 5, cy - 8), (cx, cy), (cx - 5, cy + 8)], fill="#b68b4e")
    draw.polygon([(cx + 42, cy), (cx + 5, cy - 8), (cx, cy), (cx + 5, cy + 8)], fill="#b68b4e")
    draw.text((cx - 7, cy - 75), "N", font=font_legend, fill="#442d18")

    legend_x, legend_y = width - 230, 280
    draw.rounded_rectangle([legend_x - 18, legend_y - 20, width - 42, legend_y + 520], radius=18, fill="#f3e0b8", outline="#6a4c2a", width=2)
    draw.text((legend_x, legend_y - 4), "\u56fe\u4f8b", font=font(24, True), fill="#3d2a16")
    legend_items = ["\u5f00\u59cb\u70b9", "\u4eba\u7c7b", "\u7cbe\u7075", "\u5929\u4f7f", "\u5723\u6cb3", "\u6d77\u6d0b", "\u6c99\u6f20", "\u5438\u8840\u9b3c", "\u9b54\u65cf", "\u9b54\u90fd", "\u8fdc\u53e4\u57ce\u5e02"]
    for i, item in enumerate(legend_items):
        yy = legend_y + 44 + i * 39
        draw.rounded_rectangle([legend_x, yy, legend_x + 30, yy + 24], radius=6, fill=palette.get(item, "#ddd"), outline="#5e4325")
        draw.text((legend_x + 42, yy + 1), item, font=font_legend, fill="#3d2a16")

    draw.rounded_rectangle([44, height - 78, width - 44, height - 30], radius=12, fill="#4a3520", outline="#b58a49", width=2)
    draw.text(
        (64, height - 65),
        "\u7528\u9014\u5efa\u8bae\uff1a\u4f5c\u4e3a\u7ae0\u8282\u9009\u62e9\u3001\u5267\u60c5\u767d\u5c4f\u8f6c\u573a\u6216\u4e16\u754c\u89c2\u8bf4\u660e\u9875\u7684\u5e95\u56fe\uff1b\u540e\u7eed\u53ef\u5728\u540c\u4e00\u76ee\u5f55\u66ff\u6362\u4e3a\u624b\u7ed8\u7248\u3002",
        font=font(18),
        fill="#f3dda5",
    )

    shade = Image.new("L", (width, height), 255)
    shade_draw = ImageDraw.Draw(shade)
    for i in range(105):
        shade_draw.rectangle([i, i, width - i - 1, height - i - 1], outline=max(0, 255 - i * 3))
    shade = shade.filter(ImageFilter.GaussianBlur(18))
    dark = Image.new("RGB", (width, height), "#3a2514")
    image = Image.composite(image, dark, shade)

    image.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
